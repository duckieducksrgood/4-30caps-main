from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.http import Http404, HttpResponse
from .models import Cart, Order
from .serializers import CartSerializer, OrderSerializer, OrderItemSerializer,PaymentQrSerializer
from products.models import Product
from django.shortcuts import get_object_or_404
from django.conf import settings
from users.models import User
import jwt
from products.models import Product, Category
from django.core.mail import send_mail
from .models import Order, OrderItem
from django.utils.timezone import now
from django.db.models import Q
import datetime
from .models import Order, PaymentQr
from django.utils import timezone
import datetime
import openpyxl
import pytz
from openpyxl.utils import get_column_letter
import boto3
from django.db.models import Q, Sum
from botocore.exceptions import ClientError
from rest_framework.permissions import IsAuthenticated
from django.conf import settings
from django.core.mail import send_mail, EmailMultiAlternatives
from django.utils.timezone import now
from django.shortcuts import get_object_or_404
from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView
import jwt
from products.models import Product
from users.models import User
from .models import Order, OrderItem, Cart
from .serializers import OrderSerializer, OrderItemSerializer



def create_presigned_post(bucket_name, object_name, fields=None, conditions=None, expiration=3600):
    """Generate a presigned URL S3 POST request to upload a file"""
    s3_client = boto3.client(
        's3',
        aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
        aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
        region_name=settings.AWS_S3_REGION_NAME
    )
    try:
        response = s3_client.generate_presigned_post(bucket_name, object_name, Fields=fields, Conditions=conditions, ExpiresIn=expiration)
    except ClientError as e:
        print(e)
        return None
    return response

# Add this function after all your imports but before the first class definition
def send_order_status_update_email(order, status_update, location=None, additional_info=None):
    """Helper function to send standardized order status update emails to customers"""
    user_email = order.user.email
    if not user_email:
        print(f"No email found for user: {order.user.username}")
        return False
    
    order_items = OrderItem.objects.filter(order=order)
    order_items_str = ', '.join([f"{item.product.name} (x{item.quantity})" for item in order_items])
    
    # Format estimated delivery date if available
    delivery_date_str = ""
    if order.delivery_date:
        delivery_date_str = f"Estimated delivery: {order.delivery_date.strftime('%B %d, %Y')}"
    
    # Base email subject and body
    subject = f"Order #{order.order_id} Status Update: {status_update}"
    
    # Build email body based on status
    body = f"""Dear {order.user.username},

Your order #{order.order_id} containing: {order_items_str} has been updated.

STATUS: {status_update}
"""
    
    # Add location information if provided
    if location:
        body += f"\nCURRENT LOCATION: {location}\n"
    
    # Add additional information if provided
    if additional_info:
        body += f"\nADDITIONAL INFO: {additional_info}\n"
    
    # Add delivery date if available
    if delivery_date_str:
        body += f"\n{delivery_date_str}\n"
    
    # Add status-specific messages
    if status_update == "ORDER ACCEPTED":
        body += """
Your order has been confirmed and is now being processed.
We will notify you when your order is out for delivery.
"""
    elif status_update == "PROCESSING":
        body += """
Your order is currently being prepared in our warehouse.
We'll let you know once it's ready for shipping.
"""
    elif status_update == "OUT FOR DELIVERY":
        body += """
Your order is on its way! Our delivery personnel will contact you when they're close.
Please ensure someone is available to receive the package.
"""
    elif status_update == "DELIVERED":
        body += """
Your order has been delivered successfully. Thank you for shopping with 4-30 Caps!
We hope you enjoy your purchase.

If you have any feedback or concerns, please don't hesitate to contact us.
"""
    elif status_update == "CANCELLED":
        body += """
Your order has been cancelled as requested. Any payment will be refunded according to our refund policy.
If you didn't request this cancellation, please contact our customer support immediately.
"""
    elif status_update == "REJECTED":
        body += """
We're sorry, but your order could not be processed. This could be due to:
- Payment issues
- Product availability
- Address verification problems

Please contact our customer support for assistance.
"""
    
    # Add footer
    body += """

If you need any further assistance, please contact our support team.

Thank you for choosing 4-30 Caps!
Customer Support: support@4-30caps.com
Website: www.4-30caps.com
"""
    
    # Create HTML version
    html_message = f"""
    <html>
    <head>
        <style>
            body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; }}
            .container {{ max-width: 600px; margin: 0 auto; }}
            .header {{ background-color: #4A90E2; color: white; padding: 20px; text-align: center; }}
            .content {{ padding: 20px; }}
            .footer {{ background-color: #f5f5f5; padding: 15px; text-align: center; font-size: 0.8em; }}
            .status {{ font-size: 18px; font-weight: bold; margin-bottom: 20px; }}
            .info-box {{ background-color: #e8f4fc; padding: 15px; margin: 15px 0; border-left: 4px solid #4A90E2; }}
            .location-box {{ background-color: #fff9e6; padding: 15px; margin: 15px 0; border-left: 4px solid #f1c40f; }}
            .action-btn {{ 
                display: inline-block; 
                padding: 10px 15px; 
                background-color: #4A90E2; 
                color: white !important; 
                text-decoration: none;
                border-radius: 5px;
                margin-top: 15px;
            }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <h1>Order Status Update</h1>
            </div>
            <div class="content">
                <p>Dear {order.user.username},</p>
                
                <p>Your order #{order.order_id} has been updated.</p>
                
                <div class="status">
                    Status: {status_update}
                </div>
                
                <p><strong>Items in your order:</strong> {order_items_str}</p>
    """
    
    # Add location if provided
    if location:
        html_message += f"""
                <div class="location-box">
                    <p><strong>CURRENT LOCATION:</strong> {location}</p>
                </div>
        """
    
    # Add additional info if provided
    if additional_info:
        html_message += f"""
                <div class="info-box">
                    <p><strong>ADDITIONAL INFO:</strong> {additional_info}</p>
                </div>
        """
    
    # Add delivery date if available
    if delivery_date_str:
        html_message += f"""
                <p><strong>{delivery_date_str}</strong></p>
        """
    
    # Add status-specific content
    if status_update == "ORDER ACCEPTED":
        html_message += """
                <p>Your order has been confirmed and is now being processed.</p>
                <p>We will notify you when your order is out for delivery.</p>
        """
    elif status_update == "PROCESSING":
        html_message += """
                <p>Your order is currently being prepared in our warehouse.</p>
                <p>We'll let you know once it's ready for shipping.</p>
        """
    elif status_update == "OUT FOR DELIVERY":
        html_message += """
                <p>Your order is on its way! Our delivery personnel will contact you when they're close.</p>
                <p>Please ensure someone is available to receive the package.</p>
        """
    elif status_update == "DELIVERED":
        html_message += """
                <p>Your order has been delivered successfully. Thank you for shopping with 4-30 Caps!</p>
                <p>We hope you enjoy your purchase.</p>
                <p>If you have any feedback or concerns, please don't hesitate to contact us.</p>
        """
    elif status_update == "CANCELLED":
        html_message += """
                <p>Your order has been cancelled as requested. Any payment will be refunded according to our refund policy.</p>
                <p>If you didn't request this cancellation, please contact our customer support immediately.</p>
        """
    elif status_update == "REJECTED":
        html_message += """
                <p>We're sorry, but your order could not be processed. This could be due to:</p>
                <ul>
                    <li>Payment issues</li>
                    <li>Product availability</li>
                    <li>Address verification problems</li>
                </ul>
                <p>Please contact our customer support for assistance.</p>
        """
    
    # Add footer and close HTML
    html_message += """
                <p>If you need any further assistance, please contact our support team.</p>
                
                <a href="https://4-30caps.com/track-order" class="action-btn">Track Your Order</a>
                
                <p>Thank you for choosing 4-30 Caps!</p>
            </div>
            <div class="footer">
                <p>Customer Support: support@4-30caps.com</p>
                <p>Website: <a href="https://www.4-30caps.com">www.4-30caps.com</a></p>
                <p>© 2025 4-30 Caps. All rights reserved.</p>
            </div>
        </div>
    </body>
    </html>
    """
    
    # Send the email with both HTML and plain text versions
    try:
        print(f"Attempting to send email to {user_email} with subject: {subject}")
        email = EmailMultiAlternatives(
            subject,
            body,
            settings.EMAIL_HOST_USER,
            [user_email]
        )
        email.attach_alternative(html_message, "text/html")
        email.send(fail_silently=False)
        print(f"Email sent successfully to {user_email}")
        return True
    except Exception as e:
        print(f"Error sending email notification to {user_email}: {e}")
        return False
class GeneratePresignedUrl(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        file_name = request.data.get('file_name')
        file_type = request.data.get('file_type')

        if not file_name or not file_type:
            return Response({'error': 'File name and file type are required'}, status=400)

        bucket_name = settings.AWS_STORAGE_BUCKET_NAME
        object_name = f"payment/images/{file_name}"

        presigned_post = create_presigned_post(bucket_name, object_name, fields={"Content-Type": file_type}, conditions=[{"Content-Type": file_type}])

        if presigned_post is None:
            return Response({'error': 'Could not generate presigned URL'}, status=500)

        return Response({'url': presigned_post['url'], 'fields': presigned_post['fields']}, status=200)








class ExportAnalyticsAPIView(APIView):
    def get(self, request, *args, **kwargs):
        # Get the current time in the correct timezone (Asia/Manila)
        current_time = timezone.localtime(timezone.now())

        # Get date range and type from request parameters
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        report_type = request.query_params.get('report_type', 'custom').lower()

        # Determine date range based on report type
        if report_type == 'yearly':
            start_date = current_time.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
            end_date = current_time
            period_name = f"Annual Report {start_date.year}"
        elif report_type == 'monthly':
            start_date = current_time.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            # Get last day of current month
            if current_time.month == 12:
                end_date = current_time.replace(year=current_time.year + 1, month=1, day=1) - datetime.timedelta(days=1)
            else:
                end_date = current_time.replace(month=current_time.month + 1, day=1) - datetime.timedelta(days=1)
            end_date = end_date.replace(hour=23, minute=59, second=59)
            period_name = f"Monthly Report {start_date.strftime('%B %Y')}"
        elif report_type == 'daily':
            start_date = current_time.replace(hour=0, minute=0, second=0, microsecond=0)
            end_date = current_time.replace(hour=23, minute=59, second=59)
            period_name = f"Daily Report {start_date.strftime('%B %d, %Y')}"
        else:  # custom date range
            if not start_date or not end_date:
                return Response({'message': 'Start date and end date are required for custom report'}, 
                               status=status.HTTP_400_BAD_REQUEST)
            try:
                start_date = datetime.datetime.strptime(start_date, '%Y-%m-%d')
                start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
                end_date = datetime.datetime.strptime(end_date, '%Y-%m-%d')
                end_date = end_date.replace(hour=23, minute=59, second=59)
                period_name = f"Custom Report {start_date.strftime('%b %d, %Y')} to {end_date.strftime('%b %d, %Y')}"
            except ValueError:
                return Response({'message': 'Invalid date format. Use YYYY-MM-DD'}, 
                               status=status.HTTP_400_BAD_REQUEST)

        # Create an Excel workbook
        wb = openpyxl.Workbook()
        
        # Generate Sales Report Sheet
        self._generate_sales_report(wb, start_date, end_date, period_name)
        
        # Generate Stock Report Sheet
        self._generate_stock_report(wb, period_name)

        # Save the workbook to a BytesIO object
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Create a response with the Excel file
        response = HttpResponse(
            output, 
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename={report_type}_report_{current_time.strftime("%Y%m%d")}.xlsx'
        return response

    def _generate_sales_report(self, workbook, start_date, end_date, period_name):
        """Generate the sales report worksheet"""
        # Use the active worksheet for sales report
        ws = workbook.active
        ws.title = "Sales Report"
        
        # Apply styling
        title_font = openpyxl.styles.Font(size=16, bold=True)
        header_font = openpyxl.styles.Font(size=12, bold=True)
        header_fill = openpyxl.styles.PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        
        # Add title and period
        ws['A1'] = "4-30 Caps Sales Report"
        ws['A1'].font = title_font
        ws['A2'] = period_name
        ws['A2'].font = openpyxl.styles.Font(italic=True)
        
        # Filter orders based on date range
        orders = Order.objects.filter(
            Q(delivered_date__range=[start_date, end_date]) & 
            Q(status__in=['DELIVERED', 'COMPLETED'])
        )

        # Calculate key sales metrics
        total_orders = orders.count()
        total_revenue = orders.aggregate(Sum('total_price'))['total_price__sum'] or 0
        
        # Summary statistics
        ws['A4'] = "Summary"
        ws['A4'].font = header_font
        
        summary_headers = ["Metric", "Value"]
        for col_num, header in enumerate(summary_headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f'{col_letter}5'] = header
            ws[f'{col_letter}5'].font = header_font
            ws[f'{col_letter}5'].fill = header_fill
        
        summary_metrics = [
            ["Total Completed Orders", total_orders],
            ["Total Revenue", f"₱{total_revenue:,.2f}"],
            ["Average Order Value", f"₱{(total_revenue / total_orders if total_orders else 0):,.2f}"]
        ]
        
        for row_num, metric in enumerate(summary_metrics, 6):
            ws[f'A{row_num}'] = metric[0]
            ws[f'B{row_num}'] = metric[1]
        
        # Most bought products (top 10)
        ws['A10'] = "Top Selling Products"
        ws['A10'].font = header_font
        
        product_headers = ["Product", "Category", "Quantity Sold", "Revenue"]
        for col_num, header in enumerate(product_headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f'{col_letter}11'] = header
            ws[f'{col_letter}11'].font = header_font
            ws[f'{col_letter}11'].fill = header_fill

        # Get order items for the filtered orders
        order_items = OrderItem.objects.filter(order__in=orders)
        
        # Calculate product sales
        product_sales = {}
        for item in order_items:
            product_id = item.product.productID
            if product_id not in product_sales:
                product_sales[product_id] = {
                    'name': item.product.name,
                    'category': item.product.category.name if item.product.category else 'Uncategorized',
                    'quantity': 0,
                    'revenue': 0
                }
            product_sales[product_id]['quantity'] += item.quantity
            product_sales[product_id]['revenue'] += item.quantity * item.price
        
        # Sort products by revenue (descending)
        sorted_products = sorted(
            product_sales.values(), 
            key=lambda x: x['revenue'], 
            reverse=True
        )[:10]  # Top 10 products
        
        # Add product data
        for row_num, product in enumerate(sorted_products, 12):
            ws[f'A{row_num}'] = product['name']
            ws[f'B{row_num}'] = product['category']
            ws[f'C{row_num}'] = product['quantity']
            ws[f'D{row_num}'] = f"₱{product['revenue']:,.2f}"
        
        # Daily/Monthly sales breakdown
        row_start = 12 + len(sorted_products) + 2  # Leave a gap after products list
        
        ws[f'A{row_start}'] = "Sales Breakdown by Date"
        ws[f'A{row_start}'].font = header_font
        row_start += 1
        
        # Headers for breakdown
        breakdown_headers = ["Date", "Orders", "Revenue"]
        for col_num, header in enumerate(breakdown_headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f'{col_letter}{row_start}'] = header
            ws[f'{col_letter}{row_start}'].font = header_font
            ws[f'{col_letter}{row_start}'].fill = header_fill
        
        # Group by day or month based on report length
        date_diff = (end_date - start_date).days
        
        if date_diff > 60:  # Group by month for longer periods
            breakdown = {}
            for order in orders:
                if order.delivered_date:
                    month_key = order.delivered_date.strftime('%Y-%m')
                    if month_key not in breakdown:
                        breakdown[month_key] = {'count': 0, 'revenue': 0}
                    breakdown[month_key]['count'] += 1
                    breakdown[month_key]['revenue'] += order.total_price
            
            # Sort by date
            sorted_breakdown = sorted(breakdown.items())
            
            # Add to worksheet
            for i, (month_key, data) in enumerate(sorted_breakdown, row_start + 1):
                year, month = month_key.split('-')
                month_name = datetime.date(int(year), int(month), 1).strftime('%B %Y')
                ws[f'A{i}'] = month_name
                ws[f'B{i}'] = data['count']
                ws[f'C{i}'] = f"₱{data['revenue']:,.2f}"
        else:  # Group by day for shorter periods
            breakdown = {}
            for order in orders:
                if order.delivered_date:
                    day_key = order.delivered_date.strftime('%Y-%m-%d')
                    if day_key not in breakdown:
                        breakdown[day_key] = {'count': 0, 'revenue': 0}
                    breakdown[day_key]['count'] += 1
                    breakdown[day_key]['revenue'] += order.total_price
            
            # Sort by date
            sorted_breakdown = sorted(breakdown.items())
            
            # Add to worksheet
            for i, (day_key, data) in enumerate(sorted_breakdown, row_start + 1):
                date_obj = datetime.datetime.strptime(day_key, '%Y-%m-%d')
                formatted_date = date_obj.strftime('%B %d, %Y')
                ws[f'A{i}'] = formatted_date
                ws[f'B{i}'] = data['count']
                ws[f'C{i}'] = f"₱{data['revenue']:,.2f}"
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _generate_stock_report(self, workbook, period_name):
        """Generate the stock report worksheet"""
        # Create a new worksheet for stock report
        ws = workbook.create_sheet(title="Stock Report")
        
        # Apply styling
        title_font = openpyxl.styles.Font(size=16, bold=True)
        header_font = openpyxl.styles.Font(size=12, bold=True)
        header_fill = openpyxl.styles.PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        alert_fill = openpyxl.styles.PatternFill(start_color="FFD9D9", end_color="FFD9D9", fill_type="solid")
        warning_fill = openpyxl.styles.PatternFill(start_color="FFEECC", end_color="FFEECC", fill_type="solid")
        
        # Add title and period
        ws['A1'] = "Charles Construction Services Caps Stock Report"
        ws['A1'].font = title_font
        ws['A2'] = period_name
        ws['A2'].font = openpyxl.styles.Font(italic=True)
        
        # Get all products
        all_products = Product.objects.all().order_by('category__name', 'name')
        
        # Calculate stock metrics
        total_products = all_products.count()
        low_stock_count = all_products.filter(stock__lte=5).count()
        out_of_stock_count = all_products.filter(stock=0).count()
        
        # Add summary section
        ws['A4'] = "Stock Summary"
        ws['A4'].font = header_font
        
        summary_metrics = [
            ["Total Products", total_products],
            ["Low Stock Items (<= 5)", low_stock_count],
            ["Out of Stock Items", out_of_stock_count]
        ]
        
        for row_num, metric in enumerate(summary_metrics, 5):
            ws[f'A{row_num}'] = metric[0]
            ws[f'B{row_num}'] = metric[1]
        
        # Add detailed product stock information
        ws['A8'] = "Product Stock Details"
        ws['A8'].font = header_font
        
        headers = ["Product ID", "Product Name", "Category", "Current Stock", "Status"]
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f'{col_letter}9'] = header
            ws[f'{col_letter}9'].font = header_font
            ws[f'{col_letter}9'].fill = header_fill
        
        # Group products by category
        products_by_category = {}
        for product in all_products:
            category_name = product.category.name if product.category else "Uncategorized"
            if category_name not in products_by_category:
                products_by_category[category_name] = []
            
            # Determine stock status
            if product.stock == 0:
                status = "Out of Stock"
            elif product.stock <= 5:
                status = "Low Stock"
            else:
                status = "In Stock"
                
            products_by_category[category_name].append({
                'id': product.productID,
                'name': product.name,
                'stock': product.stock,
                'status': status
            })
        
        # Add product data by category
        row_num = 10
        for category, products in sorted(products_by_category.items()):
            for product in products:
                ws[f'A{row_num}'] = product['id']
                ws[f'B{row_num}'] = product['name']
                ws[f'C{row_num}'] = category
                ws[f'D{row_num}'] = product['stock']
                ws[f'E{row_num}'] = product['status']
                
                # Apply conditional formatting based on stock level
                if product['status'] == "Out of Stock":
                    for col in range(1, 6):
                        ws.cell(row=row_num, column=col).fill = alert_fill
                elif product['status'] == "Low Stock":
                    for col in range(1, 6):
                        ws.cell(row=row_num, column=col).fill = warning_fill
                        
                row_num += 1
        
        # Add a low stock items section for quick reference
        row_num += 2
        ws[f'A{row_num}'] = "Low Stock Items Requiring Attention"
        ws[f'A{row_num}'].font = header_font
        row_num += 1
        
        # Reuse headers
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f'{col_letter}{row_num}'] = header
            ws[f'{col_letter}{row_num}'].font = header_font
            ws[f'{col_letter}{row_num}'].fill = header_fill
        
        row_num += 1
        
        # Add only low stock items
        low_stock_products = all_products.filter(stock__lte=5).order_by('stock')
        for product in low_stock_products:
            category_name = product.category.name if product.category else "Uncategorized"
            
            # Determine stock status
            if product.stock == 0:
                status = "Out of Stock"
            else:
                status = "Low Stock"
                
            ws[f'A{row_num}'] = product.productID
            ws[f'B{row_num}'] = product.name
            ws[f'C{row_num}'] = category_name
            ws[f'D{row_num}'] = product.stock
            ws[f'E{row_num}'] = status
            
            # Apply conditional formatting based on stock level
            if status == "Out of Stock":
                for col in range(1, 6):
                    ws.cell(row=row_num, column=col).fill = alert_fill
            else:
                for col in range(1, 6):
                    ws.cell(row=row_num, column=col).fill = warning_fill
                    
            row_num += 1
            
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
            
            

class CrudCart(APIView):
    def get(self, request, format=None):
        cart_user_token = request.COOKIES.get('jwt_access_token', None)
        if not cart_user_token:
            return Response({'message': 'Cart user is required'}, status=status.HTTP_401_UNAUTHORIZED)
        
        try:
            decoded_token = jwt.decode(cart_user_token, settings.SECRET_KEY, algorithms=["HS256"])
            cart_user = decoded_token.get('username')
        except jwt.ExpiredSignatureError:
            return Response({'message': 'Token has expired'}, status=status.HTTP_401_UNAUTHORIZED)
        except jwt.InvalidTokenError:
            return Response({'message': 'Invalid token'}, status=status.HTTP_401_UNAUTHORIZED)
        if not cart_user:
            return Response({'message': 'Cart user is required'}, status=status.HTTP_401_UNAUTHORIZED)
        
        
        cart_items = Cart.objects.filter(cart_user=cart_user)
        if not cart_items.exists():
            return Response({'message': 'No items found for this cart user'}, status=status.HTTP_404_NOT_FOUND)
        
        serializer = CartSerializer(cart_items, many=True)
        return Response(serializer.data)

    # pang add to cart at naauupdate ang quantity
    def post(self, request, format=None):
        cart_user_token = request.COOKIES.get('jwt_access_token', None)

        if not cart_user_token:
            return Response({'message': 'Cart user is required'}, status=status.HTTP_401_UNAUTHORIZED)
        
        try:
            decoded_token = jwt.decode(cart_user_token, settings.SECRET_KEY, algorithms=["HS256"])
            cart_user = decoded_token.get('username')
        except jwt.ExpiredSignatureError:
            return Response({'message': 'Token has expired'}, status=status.HTTP_401_UNAUTHORIZED)
        except jwt.InvalidTokenError:
            return Response({'message': 'Invalid token'}, status=status.HTTP_401_UNAUTHORIZED)
        product = request.data['product']
        quantity = request.data['quantity']

        if not cart_user:
            return Response({'message': 'Cart user is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        if product is None or quantity is None:
            return Response({'message': 'Product and quantity are required'}, status=status.HTTP_400_BAD_REQUEST)

        if not isinstance(quantity, list):
            quantity = [quantity]

        if not isinstance(product, list):
            product = [product]

        if len(quantity) != len(product):
            return Response({'message': 'Product and quantity must be of the same length'}, status=status.HTTP_400_BAD_REQUEST)

        
        # Retrieve the user instance using the cart_user
        cart_user = get_object_or_404(User, username=cart_user)

        # Loop through products and quantities
        for i in range(len(product)):
            productsCart = product[i]
            quantityCart = quantity[i]
            
            # Get the product item
            productItem = get_object_or_404(Product, productID=productsCart)

            # Get the cart price directly
            cart_price = productItem.price  # Now this is a single value, not a queryset

            cart_item_image = productItem.productImage

            # Check stock availability
            if productItem.stock < quantityCart:
                return Response({'message': 'Insufficient stock'}, status=status.HTTP_400_BAD_REQUEST)

            # Use the correct field name for cart_user
            cart, created = Cart.objects.update_or_create(
                cart_user=cart_user,  # Corrected the field name
                cart_items=productItem,
                defaults={
                    'cart_items_quantity': quantityCart,
                    'cart_price': cart_price, # Ensure this is set correctly
                    'cart_item_image': cart_item_image
                }
            )

        return Response({'message': 'Cart items added successfully'}, status=status.HTTP_201_CREATED)




    def delete(self, request, format=None):
        cart_user_token = request.COOKIES.get('jwt_access_token', None)
        if not cart_user_token:
            return Response({'message': 'Cart user is required'}, status=status.HTTP_401_UNAUTHORIZED)
        
        try:
            decoded_token = jwt.decode(cart_user_token, settings.SECRET_KEY, algorithms=["HS256"])
            cart_user = decoded_token.get('username')
        except jwt.ExpiredSignatureError:
            return Response({'message': 'Token has expired'}, status=status.HTTP_401_UNAUTHORIZED)
        except jwt.InvalidTokenError:
            return Response({'message': 'Invalid token'}, status=status.HTTP_401_UNAUTHORIZED)
        if not cart_user:
            return Response({'message': 'Cart user is required'}, status=status.HTTP_401_UNAUTHORIZED)
        
        cart_id = request.data.get('cart_id', None)

        if not cart_id:
            return Response({'message': 'Cart ID is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        cart = get_object_or_404(Cart, cart_id=cart_id, cart_user__username=cart_user)
        
        cart.delete()
        return Response({'message': 'Cart deleted successfully'}, status=status.HTTP_200_OK)

class CrudOrder(APIView):
    def get(self, request, format=None):
        order_user_token = request.COOKIES.get('jwt_access_token', None)
        if not order_user_token:
            return Response({'message': 'Order user is required'}, status=status.HTTP_401_UNAUTHORIZED)
        
        try:
            decoded_token = jwt.decode(order_user_token, settings.SECRET_KEY, algorithms=["HS256"])
            order_user = decoded_token.get('username')
        except jwt.ExpiredSignatureError:
            return Response({'message': 'Token has expired'}, status=status.HTTP_401_UNAUTHORIZED)
        except jwt.InvalidTokenError:
            return Response({'message': 'Invalid token'}, status=status.HTTP_401_UNAUTHORIZED)
        if not order_user:
            return Response({'message': 'Order user is required'}, status=status.HTTP_401_UNAUTHORIZED)
        
        orders = Order.objects.filter(user=order_user)
        order_items = OrderItem.objects.filter(order__in=orders)
        if not orders.exists():
            return Response({'message': 'No orders found for this user'}, status=status.HTTP_404_NOT_FOUND)
        
        orders_serializer = OrderSerializer(orders, many=True)
        order_items_serializer = OrderItemSerializer(order_items, many=True)
        return Response({'orders': orders_serializer.data, 'order_items': order_items_serializer.data})

    def post(self, request, format=None):
        order_user_token = request.COOKIES.get('jwt_access_token', None)
        if not order_user_token:
            return Response({'message': 'Order user is required'}, status=status.HTTP_401_UNAUTHORIZED)

        try:
            decoded_token = jwt.decode(order_user_token, settings.SECRET_KEY, algorithms=["HS256"])
            order_username = decoded_token.get('username')
        except jwt.ExpiredSignatureError:
            return Response({'message': 'Token has expired'}, status=status.HTTP_401_UNAUTHORIZED)
        except jwt.InvalidTokenError:
            return Response({'message': 'Invalid token'}, status=status.HTTP_401_UNAUTHORIZED)

        cart_ids = request.data.get('cart_id', None)
        total_price = request.data.get('total_price', None)
        delivery_date = request.data.get('delivery_date', None)
        order_status = request.data.get('status', None)
        new_delivery_address = request.data.get('delivery_address', None)
        payment_method = request.data.get('payment_method', None)
        payment_proof = request.data.get('payment_proof', None)
        

        if not order_username:
            return Response({'message': 'Order user is required'}, status=status.HTTP_400_BAD_REQUEST)

        if cart_ids is None or total_price is None:
            return Response({'message': 'Cart ID and total price are required'}, status=status.HTTP_400_BAD_REQUEST)

        if not isinstance(cart_ids, list):
            cart_ids = [cart_ids]

        try:
            order_user = User.objects.get(username=order_username)
        except User.DoesNotExist:
            return Response({'message': 'User does not exist'}, status=status.HTTP_400_BAD_REQUEST)

        # Check if the user has a delivery address
        if not order_user.delivery_address and not new_delivery_address:
            return Response({'message': 'Delivery address is required'}, status=status.HTTP_400_BAD_REQUEST)

        # Retrieve the cart items for the user
        cart_items = Cart.objects.filter(cart_id__in=cart_ids, cart_user=order_user)

        if not cart_items.exists():
            return Response({'message': 'No cart items found for the given IDs'}, status=status.HTTP_400_BAD_REQUEST)

        # Check stock availability for each cart item
        for cart_item in cart_items:
            product_stock = cart_item.cart_items.stock
            if cart_item.cart_items_quantity > product_stock:
                return Response({'message': f'Product {cart_item.cart_items.name} is out of stock'}, status=status.HTTP_400_BAD_REQUEST)

        # Use the new delivery address if provided, otherwise use the user's current delivery address
        delivery_address = new_delivery_address if new_delivery_address else order_user.delivery_address

        # Create the order and save it
        order = Order(
            user=order_user,
            total_price=total_price,
            delivery_date=delivery_date,
            status=order_status,
            order_delivery_address=delivery_address,
            payment_method=payment_method,
            payment_proof=payment_proof if payment_proof else None
        )
        order.save()  # Save the order to generate the order_id

        # Track low stock items
        low_stock_items = []

        # Deduct stock for each cart item
        for cart_item in cart_items:
            # Update product stock
            product = cart_item.cart_items
            product.stock -= cart_item.cart_items_quantity
            
            # Check if stock is low (5 or fewer items)
            if product.stock <= 5:
                low_stock_items.append({
                    'name': product.name,
                    'id': product.productID,
                    'remaining_stock': product.stock,
                    'category': product.category.name if product.category else 'Uncategorized'
                })
                
            product.save()  # Save the updated product stock

            # Create order item
            OrderItem.objects.create(
                order=order,
                product=product,
                quantity=cart_item.cart_items_quantity,
                price=cart_item.cart_price
            )

        # Delete cart items once they are checked out
        cart_items.delete()

        # Check for other low stock items across all products
        all_products = Product.objects.filter(stock__lte=5)
        
        # Add products not already in low_stock_items list
        for product in all_products:
            # Check if this product is already in the low_stock_items list
            if not any(item['id'] == product.productID for item in low_stock_items):
                low_stock_items.append({
                    'name': product.name,
                    'id': product.productID,
                    'remaining_stock': product.stock,
                    'category': product.category.name if product.category else 'Uncategorized'
                })

        # Send a detailed confirmation email to the customer
        try:
            user_email = order_user.email
            if user_email:
                print(f"Sending order confirmation to customer: {user_email}")
                
                # Get order items for the email
                order_items = OrderItem.objects.filter(order=order)
                
                # Format order details
                order_items_list = []
                for item in order_items:
                    item_price = item.price * item.quantity
                    order_items_list.append({
                        'name': item.product.name,
                        'quantity': item.quantity,
                        'price': item.price,
                        'total': item_price
                    })
                    
                # Format delivery date if available
                delivery_date_str = ""
                if delivery_date:
                    try:
                        if isinstance(delivery_date, str):
                            # Parse the string to a datetime object
                            delivery_date_obj = datetime.datetime.strptime(delivery_date, '%Y-%m-%d')
                            delivery_date_str = delivery_date_obj.strftime('%B %d, %Y')
                        else:
                            # Assume it's already a datetime object
                            delivery_date_str = delivery_date.strftime('%B %d, %Y')
                    except:
                        delivery_date_str = delivery_date  # Use as is if parsing fails
                
                # Create HTML message
                html_message = f"""
                <html>
                <head>
                    <style>
                        body {{ font-family: 'Arial', sans-serif; line-height: 1.6; color: #333; }}
                        .container {{ max-width: 600px; margin: 0 auto; }}
                        .header {{ background-color: #4A90E2; color: white; padding: 20px; text-align: center; }}
                        .content {{ padding: 20px; }}
                        .footer {{ background-color: #f5f5f5; padding: 15px; text-align: center; font-size: 0.8em; }}
                        .details {{ margin: 20px 0; }}
                        .order-number {{ font-size: 18px; font-weight: bold; color: #4A90E2; }}
                        .items-table {{ width: 100%; border-collapse: collapse; margin: 15px 0; }}
                        .items-table th {{ background-color: #f5f5f5; text-align: left; padding: 8px; }}
                        .items-table td {{ border-bottom: 1px solid #ddd; padding: 8px; }}
                        .total-row {{ font-weight: bold; }}
                        .note {{ background-color: #FFF9C4; padding: 10px; border-left: 4px solid #FFC107; margin: 15px 0; }}
                        .steps {{ background-color: #E8F5E9; padding: 10px; margin: 15px 0; }}
                        .steps h3 {{ color: #2E7D32; }}
                        .action-button {{ 
                            display: inline-block; 
                            padding: 10px 15px; 
                            background-color: #3c5a99; 
                            color: white !important;
                            text-decoration: none;
                            border-radius: 4px;
                            margin-top: 15px;
                        }}
                    </style>
                </head>
                <body>
                    <div class="container">
                        <div class="header">
                            <h1>Order Confirmation</h1>
                        </div>
                        
                        <div class="content">
                            <p>Dear {order_username},</p>
                            
                            <p>Thank you for your order with 4-30 Caps! We're excited to confirm that we've received your order.</p>
                            
                            <div class="details">
                                <p class="order-number">Order Number: #{order.order_id}</p>
                                <p><strong>Order Date:</strong> {timezone.now().strftime('%B %d, %Y')}</p>
                                <p><strong>Payment Method:</strong> {payment_method}</p>
                                <p><strong>Delivery Address:</strong> {delivery_address}</p>
                                {f"<p><strong>Estimated Delivery:</strong> {delivery_date_str}</p>" if delivery_date_str else ""}
                            </div>
                            
                            <h2>Order Items:</h2>
                            <table class="items-table">
                                <tr>
                                    <th>Item</th>
                                    <th>Quantity</th>
                                    <th>Price</th>
                                    <th>Total</th>
                                </tr>
                """
                
                # Add each item to the table
                for item in order_items_list:
                    html_message += f"""
                                <tr>
                                    <td>{item['name']}</td>
                                    <td>{item['quantity']}</td>
                                    <td>₱{item['price']:.2f}</td>
                                    <td>₱{item['total']:.2f}</td>
                                </tr>
                    """
                
                # Add total row
                html_message += f"""
                                <tr class="total-row">
                                    <td colspan="3" style="text-align: right;">Total</td>
                                    <td>₱{total_price:.2f}</td>
                                </tr>
                            </table>
                            
                            <div class="note">
                                <p><strong>Order Status: PENDING</strong></p>
                                <p>Your order is currently pending approval by our admin team. We'll review it shortly and update you once it's processed.</p>
                            </div>
                            
                            <div class="steps">
                                <h3>What happens next?</h3>
                                <ol>
                                    <li>Our team will review your order (usually within 24 hours)</li>
                                    <li>Once approved, your order will be processed and prepared for shipping</li>
                                    <li>You'll receive email updates as your order progresses</li>
                                    <li>Your 4-30 Caps items will be delivered to your address</li>
                                </ol>
                            </div>
                            
                            <p>If you have any questions about your order, please contact our customer service team at <a href="mailto:support@4-30caps.com">support@4-30caps.com</a> or call us at (123) 456-7890.</p>
                            
                            <p>Thank you for shopping with 4-30 Caps!</p>
                            
                            <p>Best Regards,<br>The 4-30 Caps Team</p>
                        </div>
                        
                        <div class="footer">
                            <p>© 2025 4-30 Caps. All rights reserved.</p>
                            <p>If you have any questions, please contact us at support@4-30caps.com</p>
                        </div>
                    </div>
                </body>
                </html>
                """
                
                # Create plain text version as fallback
                plain_message = f"""
    Order Confirmation

    Dear {order_username},

    Thank you for your order with 4-30 Caps! We're excited to confirm that we've received your order.

    ORDER DETAILS:
    Order Number: #{order.order_id}
    Order Date: {timezone.now().strftime('%B %d, %Y')}
    Payment Method: {payment_method}
    Delivery Address: {delivery_address}
    """
                
                if delivery_date_str:
                    plain_message += f"Estimated Delivery: {delivery_date_str}\n"
                    
                plain_message += """
    ORDER ITEMS:
    """
                
                for item in order_items_list:
                    plain_message += f"• {item['name']} x {item['quantity']} = ₱{item['total']:.2f}\n"
                    
                plain_message += f"""
    TOTAL: ₱{total_price:.2f}

    ORDER STATUS: PENDING
    Your order is currently pending approval by our admin team. We'll review it shortly and update you once it's processed.

    WHAT HAPPENS NEXT?
    1. Our team will review your order (usually within 24 hours)
    2. Once approved, your order will be processed and prepared for shipping
    3. You'll receive email updates as your order progresses
    4. Your 4-30 Caps items will be delivered to your address

    If you have any questions about your order, please contact our customer service team at support@4-30caps.com or call us at (123) 456-7890.

    Thank you for shopping with 4-30 Caps!

    Best Regards,
    The 4-30 Caps Team

    © 2025 4-30 Caps. All rights reserved.
    """
                
                # Send the email with both HTML and plain versions
                try:
                    email = EmailMultiAlternatives(
                        f'Your 4-30 Caps Order #{order.order_id} Confirmation',
                        plain_message,
                        settings.EMAIL_HOST_USER,
                        [user_email]
                    )
                    email.attach_alternative(html_message, "text/html")
                    email.send(fail_silently=False)  # Changed to False to catch errors
                    print(f"Successfully sent order confirmation to {user_email}")
                except Exception as e:
                    print(f"Error in EmailMultiAlternatives: {e}")
                
        except Exception as e:
            # Log the error but don't stop the process
            print(f"Error in customer email section: {e}")
            
        # Send notification to staff
        try:
            # Get all admin and employee users
            staff_users = User.objects.filter(role__in=['admin', 'employee'])
            staff_emails = [user.email for user in staff_users if user.email]
            print(f"Staff emails for notification: {staff_emails}")
            
            if staff_emails:  # Only attempt to send if there are staff emails
                # Create staff notification with more details
                html_message = f"""
                <html>
                <head>
                    <style>
                        body {{ font-family: Arial, sans-serif; }}
                        .container {{ max-width: 600px; margin: 0 auto; }}
                        .header {{ background-color: #3c5a99; color: white; padding: 15px; text-align: center; }}
                        .content {{ padding: 20px; }}
                        .order-details {{ background-color: #f5f5f5; padding: 15px; margin-bottom: 20px; }}
                        table {{ width: 100%; border-collapse: collapse; }}
                        th, td {{ padding: 8px; text-align: left; border-bottom: 1px solid #ddd; }}
                        th {{ background-color: #f2f2f2; }}
                        .action-button {{ 
                            display: inline-block; 
                            padding: 10px 15px; 
                            background-color: #3c5a99; 
                            color: white !important;
                            text-decoration: none;
                            border-radius: 4px;
                            margin-top: 15px;
                        }}
                    </style>
                </head>
                <body>
                    <div class="container">
                        <div class="header">
                            <h2>New Order Received</h2>
                        </div>
                        <div class="content">
                            <p>A new order has been placed and requires your attention.</p>
                            
                            <div class="order-details">
                                <h3>Order Details</h3>
                                <p><strong>Order ID:</strong> #{order.order_id}</p>
                                <p><strong>Customer:</strong> {order_username}</p>
                                <p><strong>Payment Method:</strong> {payment_method}</p>
                                <p><strong>Total Amount:</strong> ₱{total_price:.2f}</p>
                                <p><strong>Delivery Address:</strong> {delivery_address}</p>
                                <p><strong>Order Status:</strong> PENDING</p>
                            </div>
                            
                            <h3>Ordered Items:</h3>
                            <table>
                                <tr>
                                    <th>Product Name</th>
                                    <th>Quantity</th>
                                    <th>Price</th>
                                </tr>
                """
                
                # Add items to the HTML email
                order_items = OrderItem.objects.filter(order=order)
                for item in order_items:
                    html_message += f"""
                                <tr>
                                    <td>{item.product.name}</td>
                                    <td>{item.quantity}</td>
                                    <td>₱{item.price * item.quantity:.2f}</td>
                                </tr>
                    """
                
                # Complete the HTML email
                html_message += f"""
                            </table>
                            
                            <p>This order requires review and processing.</p>
                            <a href="https://admin.4-30caps.com/orders/{order.order_id}" class="action-button">Review Order</a>
                        </div>
                    </div>
                </body>
                </html>
                """
                
                # Create plain text version
                plain_message = f"""
    NEW ORDER RECEIVED

    A new order has been placed and requires your attention.

    ORDER DETAILS:
    Order ID: #{order.order_id}
    Customer: {order_username}
    Payment Method: {payment_method}
    Total Amount: ₱{total_price:.2f}
    Delivery Address: {delivery_address}
    Order Status: PENDING

    ORDERED ITEMS:
    """
                
                for item in order_items:
                    plain_message += f"• {item.product.name} x {item.quantity} = ₱{item.price * item.quantity:.2f}\n"
                    
                plain_message += """
    This order requires review and processing.
    Please visit the admin dashboard to review the order.

    https://admin.4-30caps.com/orders/
    """
                
                # Send the email to staff
                try:
                    email = EmailMultiAlternatives(
                        f'New Order #{order.order_id} - Action Required',
                        plain_message,
                        settings.EMAIL_HOST_USER,
                        staff_emails
                    )
                    email.attach_alternative(html_message, "text/html")
                    email.send(fail_silently=False)
                    print(f"Successfully sent staff notification to {staff_emails}")
                except Exception as e:
                    print(f"Error sending staff notification email: {e}")
                
                # If there are low stock items, send a separate alert
                if low_stock_items:
                    # Define supplier email address
                    supplier_email = "supplier@4-30caps.com"
                    
                    # Create HTML email message with mailto links
                    html_message = """
                    <html>
                    <head>
                        <style>
                            body { font-family: Arial, sans-serif; }
                            h2 { color: #d9534f; }
                            table { border-collapse: collapse; width: 100%; }
                            th, td { padding: 8px; text-align: left; border-bottom: 1px solid #ddd; color: #333333; }
                            th { background-color: #f2f2f2; }
                            .low { color: #d9534f; font-weight: bold; }
                            .button { 
                                display: inline-block; 
                                padding: 6px 12px; 
                                background-color: #007bff; 
                                color: white !important;
                                text-decoration: none;
                                border-radius: 4px;
                            }
                        </style>
                    </head>
                    <body>
                        <h2>⚠️ LOW STOCK ALERT</h2>
                        <p>The following products are low in stock and need to be reordered:</p>
                        <table>
                            <tr>
                                <th>Product</th>
                                <th>ID</th>
                                <th>Category</th>
                                <th>Remaining Stock</th>
                                <th>Action</th>
                            </tr>
                    """
                    
                    for item in low_stock_items:
                        # Create a mailto link with product info in the subject and body
                        subject = f"Reorder Request: {item['name']} (ID: {item['id']})"
                        body = f"Hello,\n\nI would like to place an order for the following product:\n\nProduct Name: {item['name']}\nProduct ID: {item['id']}\nCategory: {item['category']}\nCurrent Stock: {item['remaining_stock']}\n\nPlease arrange for delivery as soon as possible.\n\nThank you."
                        
                        # URL encode the subject and body
                        import urllib.parse
                        encoded_subject = urllib.parse.quote(subject)
                        encoded_body = urllib.parse.quote(body)
                        
                        # Create the mailto link
                        mailto_link = f"mailto:{supplier_email}?subject={encoded_subject}&body={encoded_body}"
                        
                        html_message += f"""
                            <tr>
                                <td>{item['name']}</td>
                                <td>{item['id']}</td>
                                <td>{item['category']}</td>
                                <td class="low">{item['remaining_stock']}</td>
                                <td><a href="{mailto_link}" class="button">Reorder Now</a></td>
                            </tr>
                        """
                    
                    html_message += """
                        </table>
                        <p>Please restock these items as soon as possible to avoid stockouts.</p>
                        <p>This is an automated message from your inventory management system.</p>
                    </body>
                    </html>
                    """
                    
                    # Also create a plain-text version for email clients that don't support HTML
                    plain_message = "LOW STOCK ALERT:\n\n"
                    for item in low_stock_items:
                        plain_message += f"Product: {item['name']} (ID: {item['id']})\n"
                        plain_message += f"Category: {item['category']}\n"
                        plain_message += f"Remaining Stock: {item['remaining_stock']}\n"
                        
                        # Include mailto link for plain text as well
                        subject = f"Reorder Request: {item['name']} (ID: {item['id']})"
                        body = f"Hello,\n\nI would like to place an order for the following product:\n\nProduct Name: {item['name']}\nProduct ID: {item['id']}\nCategory: {item['category']}\nCurrent Stock: {item['remaining_stock']}\n\nPlease arrange for delivery as soon as possible.\n\nThank you."
                        
                        encoded_subject = urllib.parse.quote(subject)
                        encoded_body = urllib.parse.quote(body)
                        
                        mailto_link = f"mailto:{supplier_email}?subject={encoded_subject}&body={encoded_body}"
                        plain_message += f"Reorder Link: {mailto_link}\n\n"
                    
                    plain_message += "Please restock these items as soon as possible to avoid stockouts."
                    
                    # Send the email with both HTML and plain-text versions
                    try:
                        email = EmailMultiAlternatives(
                            'LOW STOCK ALERT - Action Required',
                            plain_message,
                            settings.EMAIL_HOST_USER,
                            staff_emails
                        )
                        email.attach_alternative(html_message, "text/html")
                        email.send(fail_silently=False)
                        print("Successfully sent low stock alert")
                    except Exception as e:
                        print(f"Error sending low stock alert: {e}")
                    
        except Exception as e:
            # Log the error but don't stop the process
            print(f"Error in staff notification section: {e}")

        # Return response with the order ID after saving
        return Response({'message': 'Order placed successfully', 'order_id': order.order_id}, status=status.HTTP_201_CREATED)
    
    def put(self, request, format=None):
            order_user_token = request.COOKIES.get('jwt_access_token', None)
            if not order_user_token:
                return Response({'message': 'Order user is required'}, status=status.HTTP_401_UNAUTHORIZED)
            
            try:
                decoded_token = jwt.decode(order_user_token, settings.SECRET_KEY, algorithms=["HS256"])
                order_user = decoded_token.get('username')
                # Get the role from the token if available, otherwise assume regular user
                user_role = decoded_token.get('role', 'regular')
            except jwt.ExpiredSignatureError:
                return Response({'message': 'Token has expired'}, status=status.HTTP_401_UNAUTHORIZED)
            except jwt.InvalidTokenError:
                return Response({'message': 'Invalid token'}, status=status.HTTP_401_UNAUTHORIZED)

            order_id = request.data.get('order_id', None)
            order_status = request.data.get('status', None)
            
            # New fields for enhanced customer notifications
            current_location = request.data.get('current_location', None)
            additional_info = request.data.get('additional_info', None)
            estimated_arrival = request.data.get('estimated_arrival', None)
            
            # Field for stock adjustments (for order cancellation/rejection)
            adjust_stock = request.data.get('adjust_stock', False)

            if not order_user:
                return Response({'message': 'Order user is required'}, status=status.HTTP_400_BAD_REQUEST)
            
            if order_id is None or order_status is None:
                return Response({'message': 'Order ID and status are required'}, status=status.HTTP_400_BAD_REQUEST)
            
            order = get_object_or_404(Order, order_id=order_id)
            previous_status = order.status
            order.status = order_status.upper()
            
            # Update estimated arrival if provided
            if estimated_arrival:
                try:
                    order.delivery_date = datetime.datetime.strptime(estimated_arrival, '%Y-%m-%d')
                except ValueError:
                    pass  # Ignore if format is incorrect
            
            # Track low stock items if we're modifying stock
            low_stock_items = []
            
            # Handle stock adjustment for order cancellation/rejection
            if adjust_stock and (order_status.upper() in ['CANCELLED', 'REJECTED']) and previous_status != 'DELIVERED':
                # Get all order items
                order_items = OrderItem.objects.filter(order=order)
                
                # Return items to stock
                for item in order_items:
                    product = item.product
                    product.stock += item.quantity
                    
                    # Check if stock is still low despite returning items
                    if product.stock <= 5:
                        low_stock_items.append({
                            'name': product.name,
                            'id': product.productID,
                            'remaining_stock': product.stock,
                            'category': product.category.name if product.category else 'Uncategorized'
                        })
                        
                    product.save()
            
            order.save()

            # Debug print for tracking issue
            print(f"Sending status update to customer: {order.user.email}, Status: {order_status.upper()}")

            # Handle email notifications for status changes with enhanced information
            if previous_status != order_status.upper():
                try:
                    user_email = order.user.email
                    print(f"Customer email for notifications: {user_email}")
                    
                    if order_status.upper() == 'ACCEPTED':
                        send_order_status_update_email(
                            order, 
                            "ORDER ACCEPTED", 
                            "Processing at our warehouse",
                            additional_info
                        )
                        print(f"Sent ACCEPTED email to {user_email}")
                    
                    elif order_status.upper() == 'PROCESSING':
                        send_order_status_update_email(
                            order, 
                            "PROCESSING", 
                            "Order being prepared in our warehouse",
                            additional_info
                        )
                        print(f"Sent PROCESSING email to {user_email}")
                    
                    elif order_status.upper() == 'REJECTED':
                        send_order_status_update_email(
                            order, 
                            "REJECTED",
                            None,
                            additional_info
                        )
                        print(f"Sent REJECTED email to {user_email}")
                    
                    elif order_status.upper() == 'CANCELLED':
                        send_order_status_update_email(
                            order, 
                            "CANCELLED",
                            None,
                            additional_info
                        )
                        print(f"Sent CANCELLED email to {user_email}")
                    
                    elif order_status.upper() == 'OUT_FOR_DELIVERY':
                        send_order_status_update_email(
                            order, 
                            "OUT FOR DELIVERY",
                            current_location or "En route to delivery address",
                            f"Expected arrival: {order.delivery_date.strftime('%B %d, %Y') if order.delivery_date else 'Today'}"
                        )
                        print(f"Sent OUT_FOR_DELIVERY email to {user_email}")
                        
                    elif order_status.upper() == 'DELIVERED':
                        # Set delivered_date to current time
                        order.delivered_date = now()
                        order.save()
                        
                        send_order_status_update_email(
                            order, 
                            "DELIVERED", 
                            "Package delivered to recipient",
                            additional_info
                        )
                        print(f"Sent DELIVERED email to {user_email}")
                    
                    # For any other status change with location updates
                    elif current_location:
                        send_order_status_update_email(
                            order,
                            order_status.upper(),
                            current_location,
                            additional_info
                        )
                        print(f"Sent {order_status.upper()} with location update to {user_email}")
                except Exception as e:
                    print(f"Error sending status update email: {e}")
            
            # Send a location update even if status didn't change
            elif current_location:
                try:
                    send_order_status_update_email(
                        order,
                        "TRACKING UPDATE",
                        current_location,
                        additional_info or f"Expected delivery: {order.delivery_date.strftime('%B %d, %Y') if order.delivery_date else 'Soon'}"
                    )
                    print(f"Sent TRACKING UPDATE email to {order.user.email}")
                except Exception as e:
                    print(f"Error sending tracking update email: {e}")

            # Check for other low stock items across all products
            all_products = Product.objects.filter(stock__lte=5)
            
            # Add products not already in low_stock_items list
            for product in all_products:
                # Check if this product is already in the low_stock_items list
                if not any(item['id'] == product.productID for item in low_stock_items):
                    low_stock_items.append({
                        'name': product.name,
                        'id': product.productID,
                        'remaining_stock': product.stock,
                        'category': product.category.name if product.category else 'Uncategorized'
                    })

            # If there are low stock items, send alert to staff
            if low_stock_items:
                try:
                    # Get all admin and employee users
                    staff_users = User.objects.filter(role__in=['admin', 'employee'])
                    staff_emails = [user.email for user in staff_users if user.email]
                    
                    if staff_emails:
                        # Define supplier email address
                        supplier_email = "supplier@4-30caps.com"
                        
                        # Create HTML email message with mailto links
                        html_message = """
                        <html>
                        <head>
                            <style>
                                body { font-family: Arial, sans-serif; }
                                h2 { color: #d9534f; }
                                table { border-collapse: collapse; width: 100%; }
                                th, td { padding: 8px; text-align: left; border-bottom: 1px solid #ddd; }
                                th { background-color: #f2f2f2; }
                                .low { color: #d9534f; font-weight: bold; }
                                .button { 
                                    display: inline-block; 
                                    padding: 6px 12px; 
                                    background-color: #007bff; 
                                    color: white !important;
                                    text-decoration: none;
                                    border-radius: 4px;
                                }
                            </style>
                        </head>
                        <body>
                            <h2>⚠️ LOW STOCK ALERT</h2>
                            <p>The following products are low in stock and need to be reordered:</p>
                            <table>
                                <tr>
                                    <th>Product</th>
                                    <th>ID</th>
                                    <th>Category</th>
                                    <th>Remaining Stock</th>
                                    <th>Action</th>
                                </tr>
                        """
                        
                        for item in low_stock_items:
                            # Create a mailto link with product info in the subject and body
                            subject = f"Reorder Request: {item['name']} (ID: {item['id']})"
                            body = f"Hello,\n\nI would like to place an order for the following product:\n\nProduct Name: {item['name']}\nProduct ID: {item['id']}\nCategory: {item['category']}\nCurrent Stock: {item['remaining_stock']}\n\nPlease arrange for delivery as soon as possible.\n\nThank you."
                            
                            # URL encode the subject and body
                            import urllib.parse
                            encoded_subject = urllib.parse.quote(subject)
                            encoded_body = urllib.parse.quote(body)
                            
                            # Create the mailto link
                            mailto_link = f"mailto:{supplier_email}?subject={encoded_subject}&body={encoded_body}"
                            
                            html_message += f"""
                                <tr>
                                    <td>{item['name']}</td>
                                    <td>{item['id']}</td>
                                    <td>{item['category']}</td>
                                    <td class="low">{item['remaining_stock']}</td>
                                    <td><a href="{mailto_link}" class="button">Reorder Now</a></td>
                                </tr>
                            """
                        
                        html_message += """
                            </table>
                            <p>Please restock these items as soon as possible to avoid stockouts.</p>
                            <p>This is an automated message from your inventory management system.</p>
                        </body>
                        </html>
                        """
                        
                        # Also create a plain-text version for email clients that don't support HTML
                        plain_message = "LOW STOCK ALERT:\n\n"
                        for item in low_stock_items:
                            plain_message += f"Product: {item['name']} (ID: {item['id']})\n"
                            plain_message += f"Category: {item['category']}\n"
                            plain_message += f"Remaining Stock: {item['remaining_stock']}\n"
                            
                            # Include mailto link for plain text as well
                            subject = f"Reorder Request: {item['name']} (ID: {item['id']})"
                            body = f"Hello,\n\nI would like to place an order for the following product:\n\nProduct Name: {item['name']}\nProduct ID: {item['id']}\nCategory: {item['category']}\nCurrent Stock: {item['remaining_stock']}\n\nPlease arrange for delivery as soon as possible.\n\nThank you."
                            
                            encoded_subject = urllib.parse.quote(subject)
                            encoded_body = urllib.parse.quote(body)
                            
                            mailto_link = f"mailto:{supplier_email}?subject={encoded_subject}&body={encoded_body}"
                            plain_message += f"Reorder Link: {mailto_link}\n\n"
                        
                        plain_message += "Please restock these items as soon as possible to avoid stockouts."
                        
                        # Send the email with both HTML and plain-text versions
                        email = EmailMultiAlternatives(
                            'LOW STOCK ALERT - Action Required',
                            plain_message,
                            settings.EMAIL_HOST_USER,
                            staff_emails
                        )
                        email.attach_alternative(html_message, "text/html")
                        email.send(fail_silently=True)
                        
                except Exception as e:
                    print(f"Error sending low stock alert: {e}")

            return Response({
                'message': 'Order status updated successfully',
                'orderId': order_id,
                'status': order.status,
                'location': current_location
            }, status=status.HTTP_200_OK)    
        
class CrudPaymentQR(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, format=None):
        # Get the latest payment QR code
        qr_code = PaymentQr.objects.last()
        
        if not qr_code:
            return Response({'message': 'No QR code found'}, status=status.HTTP_404_NOT_FOUND)
        
        serializer = PaymentQrSerializer(qr_code)
        return Response(serializer.data)

    def post(self, request, format=None):
        file_name = request.data.get('file_name')

        if not file_name:
            return Response({'message': 'File name is required'}, status=status.HTTP_400_BAD_REQUEST)

        bucket_name = settings.AWS_STORAGE_BUCKET_NAME
        object_name = f"payment/images/{file_name}"

        # Delete the existing QR code if it exists
        PaymentQr.objects.all().delete()

        # Create a new QR code entry
        qr_payment = PaymentQr.objects.create(payment_qr=object_name)

        return Response({'message': 'Payment QR code uploaded successfully'}, status=status.HTTP_201_CREATED)

#admin update view
class CustomOrderAPIView(APIView):
    def post(self, request, format=None):
        order_user_token = request.COOKIES.get('jwt_access_token', None)
        if not order_user_token:
            return Response({'message': 'Order user is required'}, status=status.HTTP_401_UNAUTHORIZED)

        try:
            decoded_token = jwt.decode(order_user_token, settings.SECRET_KEY, algorithms=["HS256"])
            order_username = decoded_token.get('username')
        except jwt.ExpiredSignatureError:
            return Response({'message': 'Token has expired'}, status=status.HTTP_401_UNAUTHORIZED)
        except jwt.InvalidTokenError:
            return Response({'message': 'Invalid token'}, status=status.HTTP_401_UNAUTHORIZED)

        custom_order_description = request.data.get('custom_order_description', None)
        category_id = request.data.get('category_id', None)

        if not order_username:
            return Response({'message': 'Order user is required'}, status=status.HTTP_400_BAD_REQUEST)

        if not custom_order_description or not category_id:
            return Response({'message': 'Custom order description and category are required'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            order_user = User.objects.get(username=order_username)
        except User.DoesNotExist:
            return Response({'message': 'User does not exist'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            category = Category.objects.get(categoryID=category_id)
        except Category.DoesNotExist:
            return Response({'message': 'Invalid category'}, status=status.HTTP_400_BAD_REQUEST)

        # Create the custom order
        order = Order(
            user=order_user,
            total_price=0,  # Placeholder price
            status='PENDING',
            order_delivery_address=order_user.delivery_address,
            custom_order_description=custom_order_description,
            custom_order = True,    
        )
        order.save()  # Save the order to generate the order_id

        # Send an email notification to admin users
        admin_users = User.objects.filter(role='admin')
        admin_emails = [admin.email for admin in admin_users]

        send_mail(
            'New Custom Order Created',
            f'A new custom order with ID {order.order_id} has been created by user {order_username}.',
            'from@example.com',  
            admin_emails,
            fail_silently=False,
        )

        return Response({'message': 'Custom order placed successfully', 'order_id': order.order_id}, status=status.HTTP_201_CREATED)
    
    def put(self, request):
        order_user_token = request.COOKIES.get('jwt_access_token', None)
        if not order_user_token:
            return Response({'message': 'Order user is required'}, status=status.HTTP_401_UNAUTHORIZED)
        
        try:
            decoded_token = jwt.decode(order_user_token, settings.SECRET_KEY, algorithms=["HS256"])
            order_user = decoded_token.get('username')
        except jwt.ExpiredSignatureError:
            return Response({'message': 'Token has expired'}, status=status.HTTP_401_UNAUTHORIZED)
        except jwt.InvalidTokenError:
            return Response({'message': 'Invalid token'}, status=status.HTTP_401_UNAUTHORIZED)

        order_id = request.data.get('order_id', None)
        order_status = request.data.get('status', None)
        new_custom_price = request.data.get('custom_price', None)
        delivery_date = request.data.get('delivery_date', None)
        # image file as string for s3
        proof_of_payment = request.data.get('proof_of_payment', None)
        current_location = request.data.get('current_location', None)
        additional_info = request.data.get('additional_info', None)

        if not order_user:
            return Response({'message': 'Order user is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        if order_id is None or order_status is None:
            return Response({'message': 'Order ID and status are required'}, status=status.HTTP_400_BAD_REQUEST)
        
        order = get_object_or_404(Order, order_id=order_id)
        previous_status = order.status
        order.status = order_status.upper()
        
        if new_custom_price:
            order.total_price = new_custom_price
        
        if delivery_date:
            order.delivery_date = delivery_date
            
        order.save()

        # Handle email notifications based on status changes
        if previous_status == 'PENDING' and order_status.upper() == 'ACCEPTED':
            try:
                send_order_status_update_email(
                    order,
                    "CUSTOM ORDER ACCEPTED",
                    None,
                    f"Your custom order has been accepted with a price of ₱{new_custom_price}. The estimated delivery date is {delivery_date if delivery_date else 'to be determined'}."
                )
            except Exception as e:
                print(f"Error sending custom order acceptance email: {e}")
            
        elif previous_status == 'PENDING' and order_status.upper() == 'REJECTED':
            try:
                send_order_status_update_email(
                    order,
                    "CUSTOM ORDER REJECTED",
                    None,
                    "We're sorry, but your custom order request could not be fulfilled at this time."
                )
            except Exception as e:
                print(f"Error sending custom order rejection email: {e}")
        
        elif previous_status != order_status.upper() or current_location:
            if order_status.upper() == "OUT_FOR_DELIVERY":
                send_order_status_update_email(
                    order, 
                    "OUT FOR DELIVERY",
                    current_location or "En route to delivery address",
                    additional_info
                )
            elif order_status.upper() == "DELIVERED":
                order.delivered_date = now()
                order.save()
                
                send_order_status_update_email(
                    order, 
                    "DELIVERED", 
                    "Package delivered to recipient",
                    additional_info
                )

        if previous_status == 'ACCEPTED' and proof_of_payment:
            bucket_name = settings.AWS_STORAGE_BUCKET_NAME
            folder_name = 'payment/images/'
            file_key = f"{folder_name}/{proof_of_payment}"
            image_url = f"{file_key}"
            order.payment_proof = image_url
            order.save()
            
            # Send payment received confirmation
            try:
                send_order_status_update_email(
                    order,
                    "PAYMENT RECEIVED",
                    None,
                    f"We've received your payment for custom order #{order.order_id}. Your order is now in production."
                )
            except Exception as e:
                print(f"Error sending payment confirmation email: {e}")

        return Response({
            'message': 'Order status updated successfully',
            'orderId': order_id,
            'status': order.status
        }, status=status.HTTP_200_OK)

class AnalyticsAPIView(APIView):

    def get(self, request, *args, **kwargs):
        # Get the current time in the correct timezone (Asia/Manila)
        current_time = timezone.localtime(timezone.now())
        print("Current time: ", current_time)

        # Print current date
        print("Formatted Current Date: ", current_time.strftime('%Y-%m-%d %H:%M:%S.%f %z'))

        # Total orders
        total_orders = Order.objects.count()
        print("Total Orders: ", total_orders)

        # Start of last month
        start_of_last_month = (current_time.replace(day=1) - datetime.timedelta(days=1)).replace(day=1)
        print("Start of Last Month: ", start_of_last_month.strftime('%Y-%m-%d %H:%M:%S.%f %z'))

        # Total orders for last month
        last_month_orders = Order.objects.filter(delivered_date__lt=current_time.replace(day=1)).count()
        print("Total Orders Last Month: ", last_month_orders)

        # Last week's Monday
        last_week_monday = current_time - datetime.timedelta(days=current_time.weekday() + 7)
        print("Last Week's Monday: ", last_week_monday.strftime('%Y-%m-%d %H:%M:%S.%f %z'))

        # Total orders for last week
        start_of_last_week = last_week_monday
        end_of_last_week = start_of_last_week + datetime.timedelta(days=7)
        last_week_orders = Order.objects.filter(delivered_date__gte=start_of_last_week, delivered_date__lt=end_of_last_week).count()
        print("Total Orders Last Week: ", last_week_orders)

        # Total orders for yesterday
        yesterday = current_time - datetime.timedelta(days=1)
        print("Yesterday's Date: ", yesterday.strftime('%Y-%m-%d %H:%M:%S.%f %z'))
        deliveries_yesterday = Order.objects.filter(delivered_date__date=yesterday.date()).count()
        print("Total Orders Yesterday: ", deliveries_yesterday)

        # Total completed/delivered orders for this week
        start_of_week = current_time - datetime.timedelta(days=current_time.weekday())
        print("Start of week: ", start_of_week.strftime('%Y-%m-%d %H:%M:%S.%f %z'))
        completed_this_week = Order.objects.filter(
            Q(status__iexact='COMPLETED') | Q(status__iexact='DELIVERED'),
            delivered_date__gte=start_of_week
        ).count()

        # Total deliveries today (use timezone-aware date)
        today = current_time.date()
        print("Today's date: ", today)
        deliveries_today = Order.objects.filter(
            delivered_date__date=today
        ).count()

        # Pending orders
        pending_orders = Order.objects.filter(status__iexact='PENDING').count()

        # Monthly report (delivered, completed, cancelled)
        start_of_month = current_time.replace(day=1)
        completed_monthly = Order.objects.filter(
            status__iexact='COMPLETED', delivered_date__gte=start_of_month
        ).count()

        cancelled_monthly = Order.objects.filter(
            status__iexact='CANCELLED', delivered_date__gte=start_of_month
        ).count()

        delivered_monthly = Order.objects.filter(
            status__iexact='DELIVERED', delivered_date__gte=start_of_month
        ).count()

        all_orders = Order.objects.all()

        # Combine all the data into a response
        data = {
            "total_orders": total_orders,
            "total_orders_last_month": last_month_orders,
            "total_orders_last_week": last_week_orders,
            "total_orders_yesterday": deliveries_yesterday,
            "completed_this_week": completed_this_week,
            "deliveries_today": deliveries_today,
            "pending_orders": pending_orders,
            "monthly_report": {
                "completed": completed_monthly,
                "cancelled": cancelled_monthly,
                "delivered": delivered_monthly,
            },
            "all_orders": OrderSerializer(all_orders, many=True).data,
            "message": "Analytics data retrieved successfully"
        }

        return Response(data)


